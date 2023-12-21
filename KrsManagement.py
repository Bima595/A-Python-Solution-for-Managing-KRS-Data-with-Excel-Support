import tkinter as tk
from tkinter import ttk
from tkinter import PhotoImage, Label
from openpyxl import Workbook, load_workbook
import time



##### USER DATABASE #####
user_data = {
    "202110370311395": {"name": "SATRIA ABIMANYU P.W", "password": "395", "ipk": 3.9, "selected_courses": []},
    "202110370311393": {"name": "NABILA AZ-ZAHRO", "password": "393", "ipk": 3.3, "selected_courses": []},
    "202110370311028": {"name": "RAYHAN MUHAMMAD SAFRUDIN", "password": "028", "ipk": 3.5, "selected_courses": []},
    "202110370311161": {"name": "AHMAD RIZKY HAS", "password": "161", "ipk": 4.0, "selected_courses": []},
    "202110370311276": {"name": "FARIS EKA ISWANTO", "password": "276", "ipk": 2.3, "selected_courses": []},
    "202110370311392": {"name": "MUHAMMAD IBNU", "password": "392", "ipk": 3.3, "selected_courses": []},
    "202110370311403": {"name": "MUHAMMAD ARSYAQ FERY", "password": "393", "ipk": 3.6, "selected_courses": []},
    "202110370311405": {"name": "DEVI APRILIANTI", "password": "405", "ipk": 3.5, "selected_courses": []},
    "202110370311407": {"name": "MUHAMMAD NUR LUTHFIA QOLBA", "password": "407", "ipk": 3.7, "selected_courses": []},
    "202110370311436": {"name": "REYNALDIO AJISAKTI GOLFIRON", "password": "436", "ipk": 3.2, "selected_courses": []},
    "202110370311433": {"name": "AL GHOZI MUHAMMAD FATUR RAHMAN", "password": "433", "ipk": 3.1, "selected_courses": []},
    "202110370311434": {"name": "ANDRE ZAIDAN ABRAR", "password": "434", "ipk": 3.6, "selected_courses": []},
    "202110370311422": {"name": "SAYID MUHAMMAD ALI HISYAM FAHLEVI", "password": "422", "ipk": 4, "selected_courses": []},
    "202110370311429": {"name": "REZKY JULIANSYAH PUTRA", "password": "429", "ipk": 3.0, "selected_courses": []},
    "202110370311419": {"name": "MUHAMMAD FADHLY SYAHPUTRA", "password": "419", "ipk": 3.2, "selected_courses": []},
    "202110370311413": {"name": "ILHAM AHNAF AL HUSNA", "password": "413", "ipk": 3.9, "selected_courses": []},
    "202110370311430": {"name": "MUHAMMAD SYAHRUL NAIM", "password": "430", "ipk": 3.8, "selected_courses": []},

}

########## DATA MATA KULIAH #########
matakuliah = {
    "Pemrogrman Website": 4,
    "Kalkulus": 3,
    "Metode Penelitian": 2,
    "Pemrogaman Lanjut": 3,
    "Pemrogaman Mobile": 4,
    "Pengantar Game": 2,
    "Jaringan Komputer": 3,
    "Design Perangkat Lunak": 2,
    "Basisdata": 3,
    "Sistem Operasi": 4,
}

##### MAXIMAL SKS DEKLARASI #####
batasan_sks_ipk_35 = {
    "min_sks": 20,
    "max_sks_high_ipk": 24
}

##### INILIALISASI DATABASE #####
def load_data_from_excel():
    global user_data, selected_courses, checkbox_vars
    try:
        wb = load_workbook("KrsDatabase.xlsx")
        ws = wb.active

        for row in ws.iter_rows(min_row=2, max_col=14, max_row=ws.max_row):
            username = row[0].value
            name = row[1].value
            password = row[13].value
            stored_selected_courses = [cell.value for cell in row[2:12]]
            total_sks = row[12].value
            if username in user_data:
                user_data[username]["name"] = name
                user_data[username]["password"] = password
                user_data[username]["selected_courses"] = stored_selected_courses
            else:
                user_data[username] = {"name": name, "password": password, "ipk": 0, "selected_courses": stored_selected_courses}
                for course in stored_selected_courses:
                    checkbox_vars[course] = tk.IntVar()

    except FileNotFoundError:
        pass

    for username, data in user_data.items():
        if username not in user_data:
            user_data[username] = {"name": "", "password": data["password"], "ipk": 0, "selected_courses": []}
            for course in data["selected_courses"]:
                checkbox_vars[course] = tk.IntVar()

##### SKS UPDATE TOTAL FUNC #####
def update_status_label():
    global label_status, current_user, user_data, selected_courses
    total_sks = calculate_total_sks()
    status_text = f"Total SKS: {total_sks}"
    label_status.config(text=status_text)


##### INILIALISASI #####
root = tk.Tk()
root.geometry("1275x720+0+0")
root.resizable(0,0)
root.title("KRS Universitas Muhammadiyah Malang")
checkbox_vars = {}

user_row_counter = {}

selected_courses = []

submit_clicked = False



##### LOGIN PAGE FUNC #####
def login():
    global current_user, entry_username, entry_password, label_status, submit_clicked

    current_user = None
    submit_clicked = False

    if submit_clicked:
        return  

    username = entry_username.get()
    password = entry_password.get()

    print(f"Attempting login with username: {username}, entered password: {password}")

    if username in user_data:
        stored_password = user_data[username]["password"]
        print(f"Stored password for {username}: {stored_password}")

        if password == stored_password:
            current_user = username
            show_krs_page(user_data[username]["ipk"])
        else:
            label_status.config(text="Login gagal. Coba lagi.")
            entry_username.delete(0, tk.END)
            entry_password.delete(0, tk.END)
    else:
        # Create a new entry for the user if not found
        user_data[username] = {"name": "", "password": password, "ipk": 0, "selected_courses": []}
        current_user = username
        show_krs_page(user_data[username]["ipk"])


##### LOGIN PAGE STYLE #####
title_label = None
def login_page():
    global label_status, root, checkbox_vars, current_user, user_row_counter, selected_courses, submit_clicked, frame_login, entry_username, entry_password, title_label, subtitle_label

    title_label = tk.Label(root, text="KRS MANAGEMENT INFORMATIKA", font=("Times New Roman", 24), fg="Black")
    title_label.place(relx=0.5, rely=0.2, anchor="center")


    subtitle_label = tk.Label(root, text="UNIVERSITAS MUHAMMADIYAH MALANG", font=("Times New Roman", 24), fg="Black")
    subtitle_label.place(relx=0.5, rely=0.26, anchor="center")

    frame_login = tk.Frame(root, bg="#F5F5F5", pady=10, padx=20, bd=10, relief=tk.GROOVE)  # Tambahkan border dan relief
    frame_login.place(relx=0.5, rely=0.5, anchor="center")

    frame_login.tkraise()

    label_username = tk.Label(frame_login, text="NIM:", font=("Helvetica", 14), bg="#F5F5F5")
    label_password = tk.Label(frame_login, text="PIC:", font=("Helvetica", 14), bg="#F5F5F5")
    entry_username = tk.Entry(frame_login, font=("Helvetica", 12), bg="#FFFFFF", bd=1, relief=tk.SOLID)  # Tambahkan border dan relief
    entry_password = tk.Entry(frame_login, show="*", font=("Helvetica", 12), bg="#FFFFFF", bd=1, relief=tk.SOLID)  # Tambahkan border dan relief
    button_login = tk.Button(frame_login, text="Login", command=login, font=("Helvetica", 14), bg="#4CAF50", fg="white", bd=1, relief=tk.SOLID)  # Tambahkan border dan relief
    label_status = tk.Label(frame_login, text="", font=("Helvetica", 12), fg="red", bg="#F5F5F5")

    label_username.grid(row=1, column=0, padx=5, pady=5)
    label_password.grid(row=2, column=0, padx=5, pady=5)
    entry_username.grid(row=1, column=1, padx=5, pady=5)
    entry_password.grid(row=2, column=1, padx=5, pady=5)
    button_login.grid(row=3, column=0, columnspan=2, pady=20)
    label_status.grid(row=5, column=0, columnspan=2, pady=10)



##### KRS PAGE #####
style = ttk.Style()
row_counter = 3
calculate_total_sks = lambda: sum([matakuliah[course] for course in selected_courses])
def show_krs_page(ipk):
    global label_status, root, checkbox_vars, current_user, user_row_counter, selected_courses, submit_clicked, frame_krs, title_label, subtitle_label, row_counter

    if frame_login:
        frame_login.destroy()
        title_label.destroy()
        subtitle_label.destroy()

    for widget in root.grid_slaves():
        widget.grid_forget()

    
    ##### STYLE KRS PAGE #####
    frame_krs = tk.Frame(root, bg="#F5F5F5", pady=20, padx=20, bd=10, relief=tk.GROOVE)
    frame_krs.place(relx=0.5, rely=0.5, anchor="center")
    frame_krs.grid_columnconfigure(1, weight=1)  # Make the last column (column index 1) expandable

    title_label = ttk.Label(frame_krs, text="KRS MANAGEMENT INFORMATIKA", style="Title.TLabel")
    title_label.grid(row=0, column=0, columnspan=2)

    subtitle_label = ttk.Label(frame_krs, text="UNIVERSITAS MUHAMMADIYAH MALANG", style="Subtitle.TLabel")
    subtitle_label.grid(row=1, column=0, columnspan=2)

    label_username = ttk.Label(frame_krs, text=f"NIM: {current_user}", style="Info.TLabel")
    label_username.grid(row=2, column=0, pady=5, sticky="w")

    label_name = ttk.Label(frame_krs, text=f"NAMA: {user_data[current_user]['name']}", style="Info.TLabel")
    label_name.grid(row=2, column=1, pady=5, sticky="w")

    label_password = ttk.Label(frame_krs, text=f"PIC: {user_data[current_user]['password']}", style="Info.TLabel")
    label_password.grid(row=3, column=0, pady=5, sticky="w")

    label_status = ttk.Label(frame_krs, text="Total SKS: 0")
    label_status.grid(row=3, column=1, pady=10, sticky="e") 

    selected_courses = user_data[current_user]["selected_courses"]

    update_status_label()

    ##### KRS PAGE FUNC #####
    def get_max_sks_limit():
        global current_user, batasan_sks_ipk_35
        return batasan_sks_ipk_35["max_sks_high_ipk"] if user_data[current_user]["ipk"] > 3.5 else batasan_sks_ipk_35["min_sks"]

    def update_status():
        global selected_courses
        total_sks = calculate_total_sks()
        update_status_label()  # Call the function to update the label

    def calculate_total_sks():
        return sum(matakuliah.get(course, 0) for course in selected_courses)

    def update_selected_courses(course):
        global selected_courses, checkbox_vars

        if checkbox_vars[course].get() == 1:
            if course not in selected_courses and calculate_total_sks() + matakuliah[course] <= get_max_sks_limit():
                selected_courses.append(course)
            else:
                checkbox_vars[course].set(0)  # Uncheck the checkbox if SKS limit is exceeded
        else:
             if course in selected_courses:
                selected_courses.remove(course)
        
        selected_courses = [course for course in selected_courses if course is not None]

        update_status()

    row_counter = 3  # Start with the next row for checkbuttons

    ##### CHECK BUTTON DARI KRS #####
    for mata_kuliah, sks in matakuliah.items():
        var = tk.IntVar()
        initial_state = 1 if mata_kuliah in selected_courses else 0
        checkbutton = tk.Checkbutton(frame_krs, text=f"{mata_kuliah} - SKS: {sks}", variable=var,
                                     command=lambda m=mata_kuliah: update_selected_courses(m), onvalue=1, offvalue=0,
                                     font=("Helvetica", 12), background="#FFFFFF", bd=1, relief=tk.SOLID)
        checkbutton.select() if initial_state else checkbutton.deselect()
        checkbutton.grid(row=row_counter, column=0, columnspan=2, sticky='w', pady=5)
        checkbox_vars[mata_kuliah] = var
        row_counter += 1

        

    ##### SUBMIT BUTTON #####
    button_submit = ttk.Button(frame_krs, text="Submit", command=submit, style="Button.TButton")
    button_submit.grid(row=row_counter, column=0, pady=20)

    button_logout = ttk.Button(frame_krs, text="Logout", command=logout, style="Button.TButton")
    button_logout.grid(row=row_counter, column=1, pady=20)

    style.configure("Frame.TFrame", background="#F5F5F5", relief=tk.GROOVE)
    style.configure("Title.TLabel", font=("Times New Roman", 24), foreground="black")
    style.configure("Subtitle.TLabel", font=("Times New Roman", 18), foreground="black")
    style.configure("Info.TLabel", font=("Helvetica", 14), background="#F5F5F5")
    style.configure("Button.TButton", font=("Helvetica", 14), background="#4CAF50", borderwidth=1, relief=tk.SOLID)



##### SUBMIT FUNCTION #####
last_submission_time = 0
submission_cooldown = 60  # Adjust this value as needed
def submit():
    global user_row_counter, selected_courses, submit_clicked, last_submission_time

    current_time = time.time()

    if current_time - last_submission_time < submission_cooldown:
        label_status.config(
            text=f"Harap tunggu {int(submission_cooldown - (current_time - last_submission_time))} detik sebelum submit lagi.")
        return

    last_submission_time = current_time
    if submit_clicked:
        return  

    submit_clicked = True 

    user_data[current_user]["selected_courses"] = selected_courses

    try:
        wb = load_workbook("KrsDatabase.xlsx")
    except FileNotFoundError:
        wb = Workbook()

    ws = wb.active

    existing_row = None
    for row in ws.iter_rows(min_row=2, max_col=1, max_row=ws.max_row):
        if row[0].value == current_user:
            existing_row = row[0].row
            break

    if existing_row is not None:
        ws.cell(row=existing_row, column=2, value=user_data[current_user]["name"])  # NAMA MAHASISWA

        for i, course in enumerate(selected_courses):
            ws.cell(row=existing_row, column=i + 3, value=course)

        for i in range(len(selected_courses), 10):
            ws.cell(row=existing_row, column=i + 3, value="")

        ws.cell(row=existing_row, column=13, value=calculate_total_sks())  # Total SKS
        ws.cell(row=existing_row, column=14, value=user_data[current_user]["password"])  # PIC
        label_status.config(text="Anda Telah Berhasil Melakukan Krs Database")
    else:
        if current_user not in user_row_counter:
            user_row_counter[current_user] = ws.max_row + 1        

        data = [current_user, user_data[current_user]["name"]]  ##### MASUKIN DATA KE ROW #####

        for i, course in enumerate(selected_courses):   ##### COURSE YANG MASUK #####
            data.append(course)

        for i in range(len(selected_courses), 10):  ##### CELL KOSONG #####
            data.append("")

        data.append(calculate_total_sks())  # Total SKS
        data.append(user_data[current_user]["password"])  # PIC
        ws.append(data)
        label_status.config(text="Anda Telah Berhasil Melakukan Krs Database")
        user_row_counter[current_user] += 1

    wb.save("KrsDatabase.xlsx")


##### LOG-OUT FUNCTION #####
def logout():
    global current_user, entry_username, entry_password, label_status, submit_clicked, selected_courses
    current_user = None
    selected_courses = []
    submit_clicked = False 
    for widget in root.grid_slaves():
        widget.grid_forget()
    login_page()

    if frame_krs:
        frame_krs.destroy()

calculate_total_sks = lambda: sum(matakuliah.get(course, 0) for course in selected_courses)

##### MAIN #####    
load_data_from_excel()
login_page()
root.mainloop()